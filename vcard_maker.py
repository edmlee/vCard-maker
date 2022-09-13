from openpyxl import load_workbook


class Contact():
    def __init__(self, first_name, last_name, phone, id):
        self.first_name = first_name
        self.last_name = last_name
        self.phone = phone
        self.id = id


def clear_vcard(id):
    try:
        file = open(f"./Contacts/Contacts {id}.vcf", "w")
        file.close()
    except FileNotFoundError:
        print(f"Missing \"{folder_name}\" folder in current directory")
        quit()


def print_vcard(contact):
    file.write("BEGIN:VCARD\n")
    file.write("VERSION:2.1\n")
    file.write(f"N:{contact.last_name};{contact.first_name};;;\n")
    file.write(f"FN:{contact.first_name} {contact.last_name}\n")
    file.write(f"TEL;CELL:{contact.phone}\n")
    file.write("END:VCARD\n")


# Main
wb = load_workbook("contact_list.xlsx")
ws = wb["Contacts"]
folder_name = "Contacts" # Folder must be created in the main directory. Used to store vCards
id_col = 1 # This column will provide the name for the .vcf file
first_name_col = 2
last_name_col = 3
phone_col = 4

num_of_contacts = 14 # Number of contacts to add per vCard. Modify as required
num_of_ids = int((ws.max_row - 1) // num_of_contacts)
row = 2

for i in range(num_of_ids):
    id = int(ws.cell(row, id_col).value)
    clear_vcard(id)

    for j in range(row, row + num_of_contacts):
        try:
            with open(f"./{folder_name}/Contacts {id}.vcf", "a") as file:
                first_name = ws.cell(j, first_name_col).value
                last_name = ws.cell(j, last_name_col).value

                # Convert empty cells into blank text
                if first_name is None:
                    first_name = ""
                if last_name is None:
                    last_name = ""
                
                # Convert invalid phone number into blank text
                try:
                    phone = int(ws.cell(j, phone_col).value)
                except TypeError:
                    phone = ""

                contact = Contact(first_name, last_name, phone, id)
                print_vcard(contact)
        except FileNotFoundError:
            print(f"Missing \"{folder_name}\" folder in current directory")
            quit()
    row += num_of_contacts
